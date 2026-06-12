"""Modèles de saisie Excel et import des dossiers candidats.

Le classeur généré contient :
- « Candidats »  : une ligne par candidat ; colonnes générées depuis la grille
  (listes déroulantes pour rangs, départements, Oui/Non) ;
- « Activites »  : format long, une ligne par élément compté (publications,
  communications…) avec quantité, position d'auteur, date, porteur, bonus ;
- « Historique » : mobilités antérieures (pénalités et fenêtre « après dernier
  bénéfice ») ;
- « Listes »     : sources des listes déroulantes ;
- « Referentiel »: documentation de la grille pour l'opérateur (points, plafonds,
  où saisir chaque critère).

Le même « plan de colonnes » sert à générer le modèle et à réimporter le fichier
rempli, ce qui garantit la cohérence des deux sens.
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUI_NON = ["Oui", "Non"]
ACTIVITY_HEADERS = [
    "candidat_id",
    "element",
    "quantite",
    "position_auteur",
    "date (AAAA-MM-JJ)",
    "doi",
    "url",
    "porteur_nb",
    "bonus_nb",
]
HISTORY_HEADERS = ["candidat_id", "date_mobilite (AAAA-MM-JJ)", "date_cloture_plateforme (AAAA-MM-JJ)"]

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


# ---------------------------------------------------------------------------
# Plan de colonnes
# ---------------------------------------------------------------------------


def _count_routed_to_activities(criterion: dict) -> bool:
    """Un critère `count` va dans la feuille Activites si la saisie exige une
    granularité par élément (date, position d'auteur, bonus) ; sinon une simple
    colonne quantité dans Candidats suffit. La grille peut forcer la saisie simple
    via `excel.saisie_simple` (ex. citations : un nombre + l'URL du profil)."""
    if criterion.get("excel", {}).get("saisie_simple"):
        return False
    if criterion.get("window") == "after_last_benefit":
        return True
    if "author_position_weighting" in criterion.get("rules", []):
        return True
    if criterion.get("bonuses"):
        return True
    for item in criterion.get("items", []):
        if item.get("bonus") or item.get("shared_cap"):
            return True
    return False


def column_plan(grid: dict) -> tuple[list[dict], list[dict]]:
    """Retourne (colonnes de la feuille Candidats, éléments de la feuille Activites).

    Chaque colonne : {header, kind, criterion?, item?, index?}.
    Chaque élément d'activité : {criterion, item, key} où key = "critère :: item".
    """
    columns: list[dict] = [
        {"header": "id", "kind": "id"},
        {"header": "nom_prenom", "kind": "info", "field": "nom"},
        {"header": "population", "kind": "population"},
        {"header": "departement", "kind": "departement"},
        {"header": "pays_destination", "kind": "mobilite", "field": "pays"},
        {"header": "duree_jours", "kind": "mobilite_num", "field": "duree_jours"},
    ]
    activities: list[dict] = []

    for criterion in grid["criteria"]:
        cid = criterion["id"]
        ctype = criterion.get("type")
        if ctype == "enum":
            columns.append({"header": cid, "kind": "enum", "criterion": criterion})
            if any(o.get("bonus") for o in criterion.get("options", [])):
                columns.append(
                    {"header": f"{cid} (bonus Oui/Non)", "kind": "enum_bonus", "criterion": criterion}
                )
        elif ctype == "fixed":
            if criterion.get("is_cap"):
                columns.append({"header": f"{cid} (points)", "kind": "fixed_cap", "criterion": criterion})
            else:
                columns.append({"header": f"{cid} (Oui/Non)", "kind": "fixed", "criterion": criterion})
                for index, _bonus in enumerate(criterion.get("bonuses", [])):
                    columns.append(
                        {
                            "header": f"{cid} (bonus {index + 1} Oui/Non)",
                            "kind": "fixed_bonus",
                            "criterion": criterion,
                            "index": index,
                        }
                    )
        elif ctype == "capped":
            columns.append({"header": f"{cid} (points)", "kind": "capped", "criterion": criterion})
        elif ctype == "manual_scores":
            for item in criterion.get("items", []):
                columns.append(
                    {
                        "header": f"{cid}.{item['id']}",
                        "kind": "manual_item",
                        "criterion": criterion,
                        "item": item,
                    }
                )
        elif ctype == "formula":
            columns.append({"header": f"{cid} (n manuel)", "kind": "formula_n", "criterion": criterion})
            if "N" in criterion.get("formula", ""):
                columns.append({"header": f"{cid} (N)", "kind": "formula_N", "criterion": criterion})
        elif ctype == "count":
            if _count_routed_to_activities(criterion):
                for item in criterion.get("items", []):
                    activities.append(
                        {
                            "criterion": criterion,
                            "item": item,
                            "key": f"{cid} :: {item['id']}",
                        }
                    )
            else:
                for item in criterion.get("items", []):
                    columns.append(
                        {
                            "header": f"{cid}.{item['id']} (qte)",
                            "kind": "count_simple",
                            "criterion": criterion,
                            "item": item,
                        }
                    )
                if criterion.get("excel", {}).get("url_profil"):
                    columns.append(
                        {
                            "header": f"{cid} (url profil)",
                            "kind": "count_url",
                            "criterion": criterion,
                            "item": criterion["items"][0],
                        }
                    )

    # Bloc budgétaire en fin de feuille : montant d'indemnité (pré-rempli depuis
    # Odoo, le moteur recalcule depuis data/costs), billet estimé, frais divers.
    columns.extend(
        [
            {"header": "montant_indemnite (DA)", "kind": "mobilite_num", "field": "indemnite_da"},
            {"header": "billet_estime (DA)", "kind": "mobilite_num", "field": "billet_estime_da"},
            {"header": "frais_divers (DA)", "kind": "mobilite_num", "field": "frais_divers_da"},
        ]
    )
    return columns, activities


# ---------------------------------------------------------------------------
# Génération du modèle
# ---------------------------------------------------------------------------


def _style_header_row(ws, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = max(14, min(32, len(str(cell.value or "")) + 2))
    ws.freeze_panes = "A2"


def _list_validation(ws, source: str, col_index: int, max_row: int = 500) -> None:
    dv = DataValidation(type="list", formula1=source, allow_blank=True, showErrorMessage=True)
    ws.add_data_validation(dv)
    letter = get_column_letter(col_index)
    dv.add(f"{letter}2:{letter}{max_row}")


def build_template(grid: dict, institution: dict | None = None) -> Workbook:
    """Construit le classeur de saisie pour une grille (et un profil d'établissement)."""
    columns, activities = column_plan(grid)
    wb = Workbook()

    # --- Feuille Listes (sources des menus déroulants) -----------------------
    listes = wb.active
    listes.title = "Listes"
    list_sources: dict[str, str] = {}

    def add_list(name: str, values: list[str]) -> None:
        col = listes.max_column + 1 if listes.max_column > 1 or listes.cell(1, 1).value else 1
        listes.cell(row=1, column=col, value=name)
        for row, value in enumerate(values, start=2):
            listes.cell(row=row, column=col, value=value)
        letter = get_column_letter(col)
        list_sources[name] = f"=Listes!${letter}$2:${letter}${len(values) + 1}"

    populations = grid.get("population", [])
    if institution and institution.get("grids", {}).get(grid["id"]):
        populations = institution["grids"][grid["id"]]
    add_list("populations", list(populations))

    departements = [d["label_fr"] for d in (institution or {}).get("departements", [])]
    if departements:
        add_list("departements", departements)
    add_list("oui_non", OUI_NON)
    if activities:
        add_list("elements", [a["key"] for a in activities])
    for column in columns:
        if column["kind"] == "enum":
            add_list(
                f"enum_{column['criterion']['id']}",
                [
                    str(o.get("label_fr") or o["value"])
                    for o in column["criterion"].get("options", [])
                ],
            )
    _style_header_row(listes, listes.max_column)

    # --- Feuille Candidats ----------------------------------------------------
    cand = wb.create_sheet("Candidats")
    for index, column in enumerate(columns, start=1):
        cand.cell(row=1, column=index, value=column["header"])
        kind = column["kind"]
        if kind == "population":
            _list_validation(cand, list_sources["populations"], index)
        elif kind == "departement" and "departements" in list_sources:
            _list_validation(cand, list_sources["departements"], index)
        elif kind == "enum":
            _list_validation(cand, list_sources[f"enum_{column['criterion']['id']}"], index)
        elif kind in ("enum_bonus", "fixed", "fixed_bonus"):
            _list_validation(cand, list_sources["oui_non"], index)
    _style_header_row(cand, len(columns))

    # --- Feuille Activites ------------------------------------------------------
    act = wb.create_sheet("Activites")
    for index, header in enumerate(ACTIVITY_HEADERS, start=1):
        act.cell(row=1, column=index, value=header)
    if activities:
        _list_validation(act, list_sources["elements"], 2, max_row=2000)
    _style_header_row(act, len(ACTIVITY_HEADERS))
    act.column_dimensions["B"].width = 52
    act.column_dimensions["F"].width = 26  # doi
    act.column_dimensions["G"].width = 40  # url

    # --- Feuille Historique -----------------------------------------------------
    hist = wb.create_sheet("Historique")
    for index, header in enumerate(HISTORY_HEADERS, start=1):
        hist.cell(row=1, column=index, value=header)
    _style_header_row(hist, len(HISTORY_HEADERS))

    # --- Feuille Referentiel ------------------------------------------------------
    ref = wb.create_sheet("Referentiel")
    ref.append(["critère", "libellé", "type", "où saisir", "détail des points"])
    for criterion in grid["criteria"]:
        ctype = criterion.get("type")
        if ctype == "count":
            location = "Activites" if _count_routed_to_activities(criterion) else "Candidats"
        elif ctype == "formula":
            location = "auto (Historique) — colonne n pour forcer"
        else:
            location = "Candidats"
        ref.append(
            [
                criterion["id"],
                criterion.get("label_fr", ""),
                ctype,
                location,
                _points_summary(criterion),
            ]
        )
    _style_header_row(ref, 5)
    ref.column_dimensions["B"].width = 60
    ref.column_dimensions["E"].width = 70
    for row in ref.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.active = wb["Candidats"]
    return wb


def _points_summary(criterion: dict) -> str:
    ctype = criterion.get("type")
    if ctype == "enum":
        parts = []
        for option in criterion.get("options", []):
            text = f"{option.get('label_fr') or option['value']}={option.get('points')}"
            if option.get("bonus"):
                bonus = option["bonus"]
                text += f" (+{bonus['points']} : {bonus.get('condition_fr', 'bonus')})"
            parts.append(text)
        return " ; ".join(parts)
    if ctype == "fixed":
        text = f"{criterion.get('points')} pts"
        for bonus in criterion.get("bonuses", []):
            text += f" ; bonus +{bonus['points']} ({bonus.get('condition', '')})"
        return text
    if ctype == "capped":
        return f"plafond {criterion.get('cap_points')} pts"
    if ctype == "manual_scores":
        return " ; ".join(f"{i['id']}≤{i.get('cap_points')}" for i in criterion.get("items", []))
    if ctype == "formula":
        return f"formule {criterion.get('formula')}"
    if ctype == "count":
        parts = []
        for item in criterion.get("items", []):
            text = f"{item['id']}={item.get('points_per_unit')}/u"
            if item.get("cap_units") is not None:
                text += f" (max {item['cap_units']} u)"
            if item.get("cap_points") is not None:
                text += f" (max {item['cap_points']} pts)"
            parts.append(text)
        return " ; ".join(parts)
    return ""


def write_template(path: str | Path, grid: dict, institution: dict | None = None) -> None:
    build_template(grid, institution).save(str(path))


def write_candidates(
    path: str | Path,
    grid: dict,
    institution: dict | None,
    candidates: list[dict],
) -> None:
    """Écrit des dossiers candidats dans un classeur au format du modèle de saisie
    (inverse de read_candidates) : Candidats, Activites et Historique remplis.
    Sert aux fiches de déclaration individuelles et à la consolidation."""
    wb = build_template(grid, institution)
    cand_ws, act_ws, hist_ws = wb["Candidats"], wb["Activites"], wb["Historique"]
    columns, activities = column_plan(grid)
    act_keys = {(a["criterion"]["id"], a["item"]["id"]): a["key"] for a in activities}
    dept_labels = {
        d["id"]: d.get("label_fr", d["id"])
        for d in (institution or {}).get("departements", [])
    }

    act_row, hist_row = 2, 2
    for row, candidate in enumerate(candidates, start=2):
        entries = candidate.get("entries", {})
        mobilite = candidate.get("mobilite", {}) or {}
        for index, col in enumerate(columns, start=1):
            kind = col["kind"]
            value = None
            if kind == "id":
                value = candidate.get("id")
            elif kind == "info":
                value = candidate.get(col["field"])
            elif kind == "population":
                value = candidate.get("population")
            elif kind == "departement":
                dept = candidate.get("grouping", {}).get("departement")
                value = dept_labels.get(dept, dept)
            elif kind in ("mobilite", "mobilite_num"):
                value = mobilite.get(col["field"])
            elif kind == "enum":
                raw = entries.get(col["criterion"]["id"], {}).get("value")
                if raw is not None:
                    option = next(
                        (o for o in col["criterion"].get("options", [])
                         if str(o["value"]) == str(raw)),
                        None,
                    )
                    value = (option or {}).get("label_fr") or raw
            elif kind == "enum_bonus":
                value = "Oui" if entries.get(col["criterion"]["id"], {}).get("option_bonus") else None
            elif kind == "fixed":
                value = "Oui" if entries.get(col["criterion"]["id"], {}).get("applies") else None
            elif kind == "fixed_bonus":
                bonuses = entries.get(col["criterion"]["id"], {}).get("bonuses", [])
                value = "Oui" if col["index"] in bonuses else None
            elif kind in ("fixed_cap", "capped"):
                value = entries.get(col["criterion"]["id"], {}).get("points")
            elif kind == "manual_item":
                value = entries.get(col["criterion"]["id"], {}).get("scores", {}).get(col["item"]["id"])
            elif kind == "formula_n":
                value = entries.get(col["criterion"]["id"], {}).get("n")
            elif kind == "formula_N":
                value = entries.get(col["criterion"]["id"], {}).get("N")
            elif kind == "count_simple":
                items = entries.get(col["criterion"]["id"], {}).get("items", [])
                total = sum(i.get("count", 1) for i in items if i.get("item") == col["item"]["id"])
                value = total or None
            elif kind == "count_url":
                items = entries.get(col["criterion"]["id"], {}).get("items", [])
                with_url = next(
                    (i for i in items if i.get("item") == col["item"]["id"] and i.get("url")),
                    None,
                )
                value = (with_url or {}).get("url")
            if value is not None:
                cand_ws.cell(row=row, column=index, value=value)

        for criterion_id, entry in entries.items():
            for item in entry.get("items", []) if isinstance(entry, dict) else []:
                key = act_keys.get((criterion_id, item.get("item")))
                if key is None:
                    continue  # critère en saisie simple : déjà dans Candidats
                values = [
                    str(candidate.get("id", "")),
                    key,
                    item.get("count", 1),
                    item.get("author_position"),
                    item.get("date"),
                    item.get("doi"),
                    item.get("url"),
                    item.get("leader_count"),
                    item.get("bonus_count"),
                ]
                for col_index, cell_value in enumerate(values, start=1):
                    if cell_value is not None:
                        act_ws.cell(row=act_row, column=col_index, value=cell_value)
                act_row += 1

        for benefit in candidate.get("benefits", []):
            hist_ws.cell(row=hist_row, column=1, value=str(candidate.get("id", "")))
            if benefit.get("date"):
                hist_ws.cell(row=hist_row, column=2, value=benefit["date"])
            if benefit.get("platform_close_date"):
                hist_ws.cell(row=hist_row, column=3, value=benefit["platform_close_date"])
            hist_row += 1

    wb.save(str(path))


# ---------------------------------------------------------------------------
# Import d'un classeur rempli
# ---------------------------------------------------------------------------


def _norm_date(value) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    date.fromisoformat(text)  # lève ValueError si invalide
    return text


def _parse_bool(value) -> bool | None:
    if value is None or value == "":
        return False
    text = str(value).strip().lower()
    if text in ("oui", "yes", "1", "true", "vrai", "x"):
        return True
    if text in ("non", "no", "0", "false", "faux"):
        return False
    return None


def _parse_number(value) -> float | None:
    if value is None or value == "":
        return None
    return float(str(value).replace(",", "."))


def read_candidates(
    path: str | Path, grid: dict, institution: dict | None = None
) -> tuple[list[dict], list[str]]:
    """Lit un classeur rempli. Retourne (candidats, erreurs).

    Les erreurs référencent la feuille et la ligne ; les lignes en erreur sont
    importées au mieux (les cellules invalides sont ignorées)."""
    columns, activities = column_plan(grid)
    by_key = {a["key"]: a for a in activities}
    dept_by_label = {
        d["label_fr"]: d["id"] for d in (institution or {}).get("departements", [])
    }
    dept_ids = set(dept_by_label.values())

    wb = load_workbook(str(path), data_only=True)
    errors: list[str] = []
    candidates: dict[str, dict] = {}

    # --- Candidats -------------------------------------------------------------
    ws = wb["Candidats"]
    headers = {str(ws.cell(row=1, column=c).value): c for c in range(1, ws.max_column + 1)}
    plan_by_header = {col["header"]: col for col in columns}
    for header in headers:
        if header not in plan_by_header and header not in (None, "None"):
            errors.append(f"Candidats : colonne inconnue ignorée : {header!r}.")

    for row in range(2, ws.max_row + 1):
        cid_col = headers.get("id")
        cid = ws.cell(row=row, column=cid_col).value if cid_col else None
        if cid is None or str(cid).strip() == "":
            continue
        cid = str(cid).strip()
        candidate: dict = {"id": cid, "grouping": {}, "benefits": [], "entries": {}}
        entries = candidate["entries"]

        for header, col_index in headers.items():
            spec = plan_by_header.get(header)
            if spec is None:
                continue
            value = ws.cell(row=row, column=col_index).value
            kind = spec["kind"]
            where = f"Candidats!L{row}"
            try:
                if kind == "info" and value not in (None, ""):
                    candidate[spec["field"]] = str(value).strip()
                elif kind == "mobilite" and value not in (None, ""):
                    candidate.setdefault("mobilite", {})[spec["field"]] = str(value).strip()
                elif kind == "mobilite_num":
                    number = _parse_number(value)
                    if number is not None:
                        candidate.setdefault("mobilite", {})[spec["field"]] = number
                elif kind == "population" and value not in (None, ""):
                    candidate["population"] = str(value).strip()
                elif kind == "departement" and value not in (None, ""):
                    text = str(value).strip()
                    dept = dept_by_label.get(text, text)
                    if dept_ids and dept not in dept_ids:
                        errors.append(f"{where} : département inconnu {text!r}.")
                    candidate["grouping"]["departement"] = dept
                elif kind == "enum" and value not in (None, ""):
                    criterion = spec["criterion"]
                    text = str(value).strip()
                    # accepte le libellé français (affiché dans le menu) ou la valeur brute
                    by_text = {}
                    for option in criterion.get("options", []):
                        by_text[str(option["value"])] = str(option["value"])
                        if option.get("label_fr"):
                            by_text[str(option["label_fr"])] = str(option["value"])
                    if text not in by_text:
                        allowed = sorted(
                            str(o.get("label_fr") or o["value"])
                            for o in criterion.get("options", [])
                        )
                        errors.append(
                            f"{where} : valeur {text!r} hors barème pour "
                            f"{criterion['id']} (attendues : {', '.join(allowed)})."
                        )
                    else:
                        entries.setdefault(criterion["id"], {})["value"] = by_text[text]
                elif kind == "enum_bonus":
                    flag = _parse_bool(value)
                    if flag is None:
                        errors.append(f"{where} : valeur Oui/Non invalide ({value!r}).")
                    elif flag:
                        entries.setdefault(spec["criterion"]["id"], {})["option_bonus"] = True
                elif kind == "fixed":
                    flag = _parse_bool(value)
                    if flag is None:
                        errors.append(f"{where} : valeur Oui/Non invalide ({value!r}).")
                    elif flag:
                        entries.setdefault(spec["criterion"]["id"], {})["applies"] = True
                elif kind == "fixed_bonus":
                    flag = _parse_bool(value)
                    if flag is None:
                        errors.append(f"{where} : valeur Oui/Non invalide ({value!r}).")
                    elif flag:
                        entry = entries.setdefault(spec["criterion"]["id"], {})
                        entry.setdefault("bonuses", []).append(spec["index"])
                elif kind in ("fixed_cap", "capped"):
                    number = _parse_number(value)
                    if number is not None:
                        entries.setdefault(spec["criterion"]["id"], {})["points"] = number
                elif kind == "manual_item":
                    number = _parse_number(value)
                    if number is not None:
                        entry = entries.setdefault(spec["criterion"]["id"], {})
                        entry.setdefault("scores", {})[spec["item"]["id"]] = number
                elif kind == "formula_n":
                    number = _parse_number(value)
                    if number is not None:
                        entries.setdefault(spec["criterion"]["id"], {})["n"] = number
                elif kind == "formula_N":
                    number = _parse_number(value)
                    if number is not None:
                        entries.setdefault(spec["criterion"]["id"], {})["N"] = number
                elif kind == "count_simple":
                    number = _parse_number(value)
                    if number:
                        entry = entries.setdefault(spec["criterion"]["id"], {})
                        entry.setdefault("items", []).append(
                            {"item": spec["item"]["id"], "count": int(number)}
                        )
                elif kind == "count_url" and value not in (None, ""):
                    entry = entries.setdefault(spec["criterion"]["id"], {})
                    items = entry.setdefault("items", [])
                    target_item = next(
                        (i for i in items if i["item"] == spec["item"]["id"]), None
                    )
                    if target_item is None:
                        target_item = {"item": spec["item"]["id"], "count": 0}
                        items.append(target_item)
                    target_item["url"] = str(value).strip()
            except ValueError:
                errors.append(f"{where} : valeur invalide pour {header!r} ({value!r}).")

        if cid in candidates:
            errors.append(f"Candidats!L{row} : identifiant en double {cid!r}, ligne ignorée.")
        else:
            candidates[cid] = candidate

    # --- Activites ---------------------------------------------------------------
    if "Activites" in wb.sheetnames:
        ws = wb["Activites"]
        # lecture par nom d'en-tête : tolère les colonnes déplacées et les anciens
        # classeurs sans colonnes doi/url
        act_headers = {
            str(ws.cell(row=1, column=c).value): c for c in range(1, ws.max_column + 1)
        }

        def _cell(row: int, header: str):
            col = act_headers.get(header)
            return ws.cell(row=row, column=col).value if col else None

        for row in range(2, ws.max_row + 1):
            cid = _cell(row, "candidat_id")
            element = _cell(row, "element")
            if cid is None and element is None:
                continue
            where = f"Activites!L{row}"
            cid = str(cid).strip() if cid is not None else ""
            if cid not in candidates:
                errors.append(f"{where} : candidat inconnu {cid!r}.")
                continue
            spec = by_key.get(str(element).strip()) if element else None
            if spec is None:
                if element in (None, ""):
                    # ligne d'aide pré-remplie (candidat_id seul) : ignorée sauf si
                    # des données ont été saisies sans choisir l'élément
                    has_data = any(
                        _cell(row, header) not in (None, "")
                        for header in ACTIVITY_HEADERS[2:]
                    )
                    if not has_data:
                        continue
                    errors.append(f"{where} : données saisies sans élément sélectionné.")
                else:
                    errors.append(f"{where} : élément inconnu {element!r}.")
                continue
            try:
                quantity = _parse_number(_cell(row, "quantite"))
                position = _parse_number(_cell(row, "position_auteur"))
                date_value = _norm_date(_cell(row, "date (AAAA-MM-JJ)"))
                leader = _parse_number(_cell(row, "porteur_nb"))
                bonus = _parse_number(_cell(row, "bonus_nb"))
            except ValueError as exc:
                errors.append(f"{where} : {exc}")
                continue
            item_entry: dict = {
                "item": spec["item"]["id"],
                "count": int(quantity) if quantity else 1,
            }
            if position is not None:
                item_entry["author_position"] = int(position)
            if date_value:
                item_entry["date"] = date_value
            doi = _cell(row, "doi")
            if doi not in (None, ""):
                item_entry["doi"] = str(doi).strip()
            url = _cell(row, "url")
            if url not in (None, ""):
                item_entry["url"] = str(url).strip()
            if leader:
                item_entry["leader_count"] = int(leader)
            if bonus:
                item_entry["bonus_count"] = int(bonus)
            entry = candidates[cid]["entries"].setdefault(spec["criterion"]["id"], {})
            entry.setdefault("items", []).append(item_entry)

    # --- Historique -----------------------------------------------------------------
    if "Historique" in wb.sheetnames:
        ws = wb["Historique"]
        for row in range(2, ws.max_row + 1):
            cid = ws.cell(row=row, column=1).value
            if cid is None or str(cid).strip() == "":
                continue
            where = f"Historique!L{row}"
            cid = str(cid).strip()
            if cid not in candidates:
                errors.append(f"{where} : candidat inconnu {cid!r}.")
                continue
            try:
                mobility = _norm_date(ws.cell(row=row, column=2).value)
                closing = _norm_date(ws.cell(row=row, column=3).value)
            except ValueError:
                errors.append(f"{where} : date invalide.")
                continue
            if mobility is None:
                errors.append(f"{where} : date de mobilité manquante.")
                continue
            benefit = {"date": mobility}
            if closing:
                benefit["platform_close_date"] = closing
            candidates[cid]["benefits"].append(benefit)

    return list(candidates.values()), errors
