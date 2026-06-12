"""Exports officiels : PV de classement et fiches d'évaluation individuelles.

Formats produits :
- Excel (.xlsx) — documents de travail à faire signer par le conseil
  scientifique / la commission (colonnes Décision et Motif à compléter,
  le rejet devant être motivé — art. 14-15 de l'arrêté 345) ;
- HTML imprimable (— PDF via « Imprimer > Enregistrer en PDF »).
"""

from __future__ import annotations

import html
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from classement.models import RankedCandidate, ScoreBreakdown

_TITLE_FONT = Font(bold=True, size=13)
_SUB_FONT = Font(bold=True, size=11)
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_THIN = Border(*(Side(style="thin"),) * 4)


def _group_title(key: tuple) -> str:
    return " / ".join(str(part) for part in key[1:])  # sans l'id de grille


def _candidate_name(candidate: dict) -> str:
    name = " ".join(filter(None, [candidate.get("nom"), candidate.get("prenom")]))
    return name or str(candidate.get("id", "?"))


def _header_block(ws, institution: dict | None, grid: dict, campaign_date: str | None, row: int = 1) -> int:
    if institution:
        ws.cell(row=row, column=1, value=institution.get("nom_fr", institution.get("id"))).font = _TITLE_FONT
        row += 1
    ws.cell(row=row, column=1, value=grid.get("title_fr", grid["id"])).font = _SUB_FONT
    row += 1
    ws.cell(
        row=row,
        column=1,
        value=f"Arrêté n° 345 du 09/03/2026 — campagne du {campaign_date or date.today().isoformat()}",
    )
    return row + 2


def _style_table_header(ws, row: int, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _THIN
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# PV de classement
# ---------------------------------------------------------------------------

PV_HEADERS = ["Rang", "ID", "Nom et prénom", "Département", "Score", "Ex aequo", "Décision", "Motif (si rejet)"]


def export_pv(
    path: str | Path,
    groups: dict[tuple, list[RankedCandidate]],
    candidates: list[dict],
    grid: dict,
    institution: dict | None = None,
    campaign_date: str | None = None,
) -> None:
    """PV de classement : une feuille par groupe (population / regroupement)."""
    by_id = {str(c.get("id")): c for c in candidates}
    wb = Workbook()
    wb.remove(wb.active)

    for index, (key, ranked) in enumerate(groups.items(), start=1):
        title = _group_title(key)
        ws = wb.create_sheet(f"{index:02d} {title}"[:31])
        row = _header_block(ws, institution, grid, campaign_date)
        ws.cell(row=row, column=1, value=f"Classement : {title}").font = _SUB_FONT
        row += 1

        header_row = row
        for col, header in enumerate(PV_HEADERS, start=1):
            ws.cell(row=header_row, column=col, value=header)
        _style_table_header(ws, header_row, len(PV_HEADERS))

        dv = DataValidation(type="list", formula1='"Accepté,Rejeté"', allow_blank=True)
        ws.add_data_validation(dv)

        for offset, entry in enumerate(ranked, start=1):
            candidate = by_id.get(entry.candidate_id, {})
            r = header_row + offset
            values = [
                entry.rank,
                entry.candidate_id,
                _candidate_name(candidate),
                candidate.get("grouping", {}).get("departement", ""),
                round(entry.total, 2),
                "oui" if entry.ex_aequo else "",
                "",
                "",
            ]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=r, column=col, value=value)
                cell.border = _THIN
            dv.add(ws.cell(row=r, column=7))

        widths = [7, 10, 30, 26, 9, 9, 12, 40]
        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width

        sig = header_row + len(ranked) + 3
        ws.cell(row=sig, column=1, value="Fait à ………………………, le ………………………")
        ws.cell(row=sig + 2, column=5, value="Le Président du Conseil Scientifique / de la Commission")

    wb.save(str(path))


# ---------------------------------------------------------------------------
# Fiches d'évaluation individuelles
# ---------------------------------------------------------------------------


def _ranks_by_candidate(groups: dict[tuple, list[RankedCandidate]]) -> dict[str, tuple[int, int, str]]:
    out: dict[str, tuple[int, int, str]] = {}
    for key, ranked in groups.items():
        for entry in ranked:
            out[entry.candidate_id] = (entry.rank, len(ranked), _group_title(key))
    return out


def export_fiches(
    path: str | Path,
    breakdowns: list[ScoreBreakdown],
    candidates: list[dict],
    groups: dict[tuple, list[RankedCandidate]],
    grid: dict,
    institution: dict | None = None,
    campaign_date: str | None = None,
) -> None:
    """Fiches d'évaluation : une feuille par candidat (détail par critère, total,
    rang, décision et motif à compléter)."""
    by_id = {str(c.get("id")): c for c in candidates}
    ranks = _ranks_by_candidate(groups)
    wb = Workbook()
    wb.remove(wb.active)

    for breakdown in breakdowns:
        candidate = by_id.get(breakdown.candidate_id, {})
        ws = wb.create_sheet(str(breakdown.candidate_id)[:31])
        row = _header_block(ws, institution, grid, campaign_date)

        ws.cell(row=row, column=1, value="Fiche d'évaluation du candidat").font = _SUB_FONT
        row += 1
        info = [
            ("Nom et prénom", _candidate_name(candidate)),
            ("Identifiant", breakdown.candidate_id),
            ("Population", breakdown.population or ""),
            ("Département", candidate.get("grouping", {}).get("departement", "")),
        ]
        for label, value in info:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1
        row += 1

        headers = ["Critère", "Points", "Détail", "Observations"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=row, column=col, value=header)
        _style_table_header(ws, row, len(headers))
        row += 1

        for line in breakdown.lines:
            if not line.points and not line.details and not line.warnings:
                continue
            cells = [
                line.label,
                round(line.points, 2),
                "\n".join(line.details),
                "\n".join(line.warnings),
            ]
            for col, value in enumerate(cells, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = _THIN
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            row += 1

        ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
        total_cell = ws.cell(row=row, column=2, value=round(breakdown.total, 2))
        total_cell.font = Font(bold=True)
        total_cell.border = _THIN
        row += 1

        rank_info = ranks.get(breakdown.candidate_id)
        if rank_info:
            rank, size, group = rank_info
            ws.cell(row=row, column=1, value="Rang").font = Font(bold=True)
            ws.cell(row=row, column=2, value=f"{rank} / {size} ({group})")
            row += 1

        for warning in breakdown.warnings:
            ws.cell(row=row, column=1, value=f"⚠ {warning}")
            row += 1
        row += 1

        ws.cell(row=row, column=1, value="Décision de la commission :").font = Font(bold=True)
        dv = DataValidation(type="list", formula1='"Accepté,Rejeté"', allow_blank=True)
        ws.add_data_validation(dv)
        decision = ws.cell(row=row, column=2)
        decision.border = _THIN
        dv.add(decision)
        row += 1
        ws.cell(row=row, column=1, value="Motif (obligatoire en cas de rejet) :").font = Font(bold=True)
        motif = ws.cell(row=row, column=2)
        motif.border = _THIN
        ws.row_dimensions[row].height = 40
        row += 2
        ws.cell(row=row, column=1, value="Signature :")

        ws.column_dimensions["A"].width = 46
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 52
        ws.column_dimensions["D"].width = 42

    wb.save(str(path))


# ---------------------------------------------------------------------------
# HTML imprimable (PV + fiches) — PDF via impression navigateur
# ---------------------------------------------------------------------------

_CSS = """
body { font-family: 'Segoe UI', Arial, sans-serif; margin: 2em; color: #111; }
h1 { font-size: 1.25em; } h2 { font-size: 1.1em; margin-top: 1.4em; }
table { border-collapse: collapse; width: 100%; margin: .6em 0 1.2em; }
th, td { border: 1px solid #555; padding: 4px 8px; font-size: .85em; vertical-align: top; }
th { background: #1F4E78; color: #fff; }
.total { font-weight: bold; background: #eef; }
.warn { color: #8a4500; font-size: .8em; }
.sig { margin-top: 2.5em; display: flex; justify-content: space-between; }
.fiche { page-break-before: always; }
@media print { body { margin: 1cm; } }
"""


def export_html(
    path: str | Path,
    groups: dict[tuple, list[RankedCandidate]],
    breakdowns: list[ScoreBreakdown],
    candidates: list[dict],
    grid: dict,
    institution: dict | None = None,
    campaign_date: str | None = None,
) -> None:
    """Document HTML auto-porté : PV par groupe puis une fiche par candidat,
    avec sauts de page pour l'impression (PDF via le navigateur)."""
    by_id = {str(c.get("id")): c for c in candidates}
    ranks = _ranks_by_candidate(groups)
    e = html.escape
    parts: list[str] = [
        "<!doctype html><html lang='fr'><head><meta charset='utf-8'>",
        f"<title>PV de classement — {e(grid.get('title_fr', grid['id']))}</title>",
        f"<style>{_CSS}</style></head><body>",
    ]
    if institution:
        parts.append(f"<h1>{e(institution.get('nom_fr', institution['id']))}</h1>")
    parts.append(f"<p><b>{e(grid.get('title_fr', grid['id']))}</b><br>")
    parts.append(
        f"Arrêté n° 345 du 09/03/2026 — campagne du {e(campaign_date or date.today().isoformat())}</p>"
    )

    for key, ranked in groups.items():
        parts.append(f"<h2>Classement : {e(_group_title(key))}</h2>")
        parts.append("<table><tr>" + "".join(f"<th>{e(h)}</th>" for h in PV_HEADERS) + "</tr>")
        for entry in ranked:
            candidate = by_id.get(entry.candidate_id, {})
            cells = [
                str(entry.rank),
                entry.candidate_id,
                _candidate_name(candidate),
                candidate.get("grouping", {}).get("departement", ""),
                f"{entry.total:g}",
                "oui" if entry.ex_aequo else "",
                "",
                "",
            ]
            parts.append("<tr>" + "".join(f"<td>{e(c)}</td>" for c in cells) + "</tr>")
        parts.append("</table>")
    parts.append(
        "<div class='sig'><span>Fait à ………………, le ………………</span>"
        "<span>Le Président du Conseil Scientifique / de la Commission</span></div>"
    )

    for breakdown in breakdowns:
        candidate = by_id.get(breakdown.candidate_id, {})
        parts.append("<div class='fiche'>")
        parts.append(f"<h2>Fiche d'évaluation — {e(_candidate_name(candidate))} ({e(breakdown.candidate_id)})</h2>")
        rank_info = ranks.get(breakdown.candidate_id)
        meta = f"Population : {e(breakdown.population or '')} — Département : " \
               f"{e(candidate.get('grouping', {}).get('departement', ''))}"
        if rank_info:
            meta += f" — Rang : {rank_info[0]} / {rank_info[1]} ({e(rank_info[2])})"
        parts.append(f"<p>{meta}</p>")
        parts.append("<table><tr><th>Critère</th><th>Points</th><th>Détail</th><th>Observations</th></tr>")
        for line in breakdown.lines:
            if not line.points and not line.details and not line.warnings:
                continue
            parts.append(
                "<tr>"
                f"<td>{e(line.label)}</td><td>{line.points:g}</td>"
                f"<td>{'<br>'.join(e(d) for d in line.details)}</td>"
                f"<td class='warn'>{'<br>'.join(e(w) for w in line.warnings)}</td>"
                "</tr>"
            )
        parts.append(
            f"<tr class='total'><td>TOTAL</td><td>{breakdown.total:g}</td><td></td><td></td></tr></table>"
        )
        for warning in breakdown.warnings:
            parts.append(f"<p class='warn'>⚠ {e(warning)}</p>")
        parts.append(
            "<p><b>Décision de la commission :</b> Accepté ☐&nbsp;&nbsp;Rejeté ☐<br>"
            "<b>Motif (obligatoire en cas de rejet) :</b> "
            "………………………………………………………………………………</p>"
            "<div class='sig'><span>Signature :</span></div></div>"
        )

    parts.append("</body></html>")
    Path(path).write_text("\n".join(parts), encoding="utf-8")
