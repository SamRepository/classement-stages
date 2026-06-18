"""Extrait l'historique des bénéfices (séjours antérieurs) d'un export Odoo.

À partir de l'export `stages.candidature.sejour_historique.xlsx` (une ligne par
séjour/candidature, toutes sessions confondues), produit une feuille **Historique**
au format attendu par l'import de comptes
(`webapp.services.accounts._import_benefits`) : colonnes ``candidat`` (référence
du dossier de la campagne en cours), ``date_mobilite`` et ``date_cloture``.

Règles (décision ENSET, première campagne) :
- on conserve les séjours **jusqu'à l'année de coupure** incluse (défaut 2025) ;
  la session en cours (2026) est exclue (pas encore bénéficié) ;
- chaque séjour 2023-2025 (état *Réalisée* ou *Confirmée*) compte comme un bénéfice ;
- ``date_mobilite`` = date de départ du séjour ; ``date_cloture`` = fin de la
  fenêtre de session (clôture de plateforme), lue dans la colonne « Session » ;
- chaque bénéfice est rattaché à la **référence 2026** du même e-mail (dossier de
  la campagne en cours). Les e-mails sans candidature 2026 sont listés en
  avertissement (historique non rattachable).

Usage :
    python scripts/extract_historique.py [--source <export.xlsx>] [--out <hist.xlsx>]
        [--cutoff-year 2025]
"""

from __future__ import annotations

import argparse
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook

DEFAULT_SOURCE = Path("examples/enset/stages.candidature.sejour_historique.xlsx")
DEFAULT_OUT = Path("examples/enset/historique-beneficiaires.xlsx")

_YEAR_RE = re.compile(r"/(\d{4})/")
_SESSION_END_RE = re.compile(r"-\s*(\d{4}-\d{2}-\d{2})\)\s*$")


def _year(ref: str) -> int | None:
    m = _YEAR_RE.search(str(ref or ""))
    return int(m.group(1)) if m else None


def _session_close(session: str) -> str | None:
    m = _SESSION_END_RE.search(str(session or ""))
    return m.group(1) if m else None


def _as_iso(value) -> str | None:
    if value is None or value == "":
        return None
    return str(value)[:10]


def extract(source: Path, cutoff_year: int) -> tuple[list[dict], list[str]]:
    """Retourne (lignes d'historique, avertissements)."""
    wb = load_workbook(source, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    col = {name: i for i, name in enumerate(rows[0])}

    def cell(row, name):
        return row[col[name]] if name in col else None

    by_email: dict[str, list] = defaultdict(list)
    for row in rows[1:]:
        email = (cell(row, "Email") or "").strip()
        if email:
            by_email[email].append(row)

    # Référence du dossier de la campagne courante (année > cutoff) par e-mail.
    current_ref: dict[str, str] = {}
    for email, candidatures in by_email.items():
        recents = [r for r in candidatures if (_year(cell(r, "Référence")) or 0) > cutoff_year]
        if recents:
            current_ref[email] = str(cell(recents[0], "Référence"))

    out: list[dict] = []
    warnings: list[str] = []
    for email in sorted(by_email):
        benefices = [
            r for r in by_email[email]
            if (_year(cell(r, "Référence")) or 9999) <= cutoff_year
        ]
        if not benefices:
            continue
        ref = current_ref.get(email, "")
        if not ref:
            warnings.append(
                f"{email} : {len(benefices)} séjour(s) jusqu'à {cutoff_year} mais aucune "
                f"candidature > {cutoff_year} — historique non rattachable à la campagne en cours."
            )
        for r in sorted(benefices, key=lambda r: _as_iso(cell(r, "Date de départ")) or ""):
            out.append({
                "candidat": ref,
                "email": email,
                "enseignant": cell(r, "Enseignant") or "",
                "date_mobilite": _as_iso(cell(r, "Date de départ")),
                "date_cloture": _session_close(cell(r, "Session")),
                "session": cell(r, "Session") or "",
                "etat": cell(r, "Etat Candidature") or "",
            })
    return out, warnings


def write_workbook(lignes: list[dict], out: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Historique"
    headers = ["candidat", "email", "enseignant", "date_mobilite", "date_cloture",
               "session", "etat"]
    ws.append(headers)
    for ligne in lignes:
        ws.append([ligne[h] for h in headers])
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)


def main(argv=None) -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT)
    parser.add_argument("--cutoff-year", type=int, default=2025,
                        help="Dernière année incluse dans l'historique (défaut 2025).")
    args = parser.parse_args(argv)

    lignes, warnings = extract(args.source, args.cutoff_year)
    write_workbook(lignes, args.out)

    par_email: dict[str, int] = defaultdict(int)
    rattaches = 0
    for ligne in lignes:
        par_email[ligne["email"]] += 1
        if ligne["candidat"]:
            rattaches += 1
    print(f"Historique écrit dans {args.out}")
    print(f"  {len(par_email)} enseignant(s), {len(lignes)} bénéfice(s) jusqu'à "
          f"{args.cutoff_year} ({rattaches} rattachés à un dossier 2026).")
    print("  n par enseignant :")
    for email in sorted(par_email):
        print(f"    {email:35} n = {par_email[email]}")
    if warnings:
        print(f"\n  {len(warnings)} avertissement(s) :")
        for w in warnings:
            print(f"   - {w}")


if __name__ == "__main__":
    main()
