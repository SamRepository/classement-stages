"""Génère une fiche de déclaration individuelle par candidat à partir du dossier
maître (ex. dossier-u3.xlsx produit par import_odoo.py).

Chaque fiche est un classeur au format du modèle de saisie, pré-rempli avec
l'identité, la destination et l'historique du candidat, plus une feuille
Instructions. Le candidat complète ses critères (feuille Candidats) et ses
activités (feuille Activites, une ligne par élément), joint ses justificatifs
PDF, et renvoie la fiche — la consolidation se fait avec
scripts/consolider_declarations.py.

Usage :
    python scripts/fiches_declaration.py --dossier examples/enset/dossier-u3.xlsx
        [--grid u3-residences-scientifiques] [--out-dir declarations]
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook
from openpyxl.styles import Font

from classement.excel_io import read_candidates, write_candidates
from classement.grids import find_grid
from classement.institutions import load_institution

INSTRUCTIONS = [
    ("Fiche de déclaration — mobilité de courte durée à l'étranger (arrêté n° 345)", True),
    ("", False),
    ("1. Vérifiez votre identité, votre destination et votre durée dans la feuille « Candidats »", False),
    ("   (signalez toute erreur au service, ne modifiez pas les en-têtes de colonnes).", False),
    ("2. Complétez vos critères dans la feuille « Candidats » : menus déroulants Oui/Non,", False),
    ("   quantités, nombre de citations + URL de votre profil Scopus ou Google Scholar,", False),
    ("   estimation du billet d'avion et des frais divers (visa, assurance) en DA.", False),
    ("3. Déclarez vos activités dans la feuille « Activites » : UNE LIGNE PAR ÉLÉMENT", False),
    ("   (publication, communication, encadrement…), avec la date, le DOI ou l'URL pour", False),
    ("   les publications, et votre position dans la liste des auteurs.", False),
    ("4. La feuille « Referentiel » détaille le barème de chaque critère.", False),
    ("5. Joignez un justificatif PDF par élément déclaré, nommé :", False),
    ("   <votre référence>_<n° de ligne Activites>.pdf (ex. DC-2026-291_L3.pdf).", False),
    ("6. Renvoyez cette fiche et vos justificatifs au service avant la date limite.", False),
    ("", False),
    ("Important : seuls les travaux postérieurs à votre dernier bénéfice de mobilité sont", False),
    ("comptabilisés (voir feuille Historique). Toute déclaration est vérifiée par la", False),
    ("commission sur pièce ; un rejet est toujours motivé (art. 14-15 de l'arrêté 345).", False),
]


def _safe_name(text: str) -> str:
    return re.sub(r"[^A-Za-z0-9_-]+", "-", str(text)).strip("-")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--dossier", required=True, help="Dossier maître (.xlsx) avec les candidats.")
    parser.add_argument("--grid", default="u3-residences-scientifiques")
    parser.add_argument("--institution", default="enset-skikda")
    parser.add_argument("--out-dir", default=str(ROOT / "declarations"))
    args = parser.parse_args()

    grid = find_grid(args.grid)
    institution = load_institution(args.institution)
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    candidates, errors = read_candidates(args.dossier, grid, institution)
    for error in errors:
        print(f"  ⚠ {error}")

    generated = 0
    for candidate in candidates:
        cid = str(candidate.get("id", "?"))
        name = _safe_name(f"{cid}_{candidate.get('nom', '')}")
        path = out_dir / f"fiche_{name}.xlsx"
        write_candidates(path, grid, institution, [candidate])

        wb = load_workbook(path)
        # pré-remplir candidat_id sur 40 lignes d'activités pour guider la saisie
        act = wb["Activites"]
        for row in range(2, 42):
            if act.cell(row=row, column=1).value in (None, ""):
                act.cell(row=row, column=1, value=cid)
        # feuille Instructions en tête
        instructions = wb.create_sheet("Instructions", 0)
        for row, (text, bold) in enumerate(INSTRUCTIONS, start=1):
            cell = instructions.cell(row=row, column=1, value=text)
            if bold:
                cell.font = Font(bold=True, size=13)
        instructions.column_dimensions["A"].width = 100
        instructions.cell(row=len(INSTRUCTIONS) + 2, column=1,
                          value=f"Candidat : {candidate.get('nom', '')} — Référence : {cid}").font = Font(bold=True)
        wb.active = wb["Instructions"]
        wb.save(path)
        generated += 1

    print(f"{generated} fiche(s) de déclaration générée(s) dans {out_dir}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
