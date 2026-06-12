"""Convertit l'export Excel du module Odoo « Stages » (stages.candidature.sejour)
en classeur de saisie pré-rempli pour la campagne u3.

Usage :
    python scripts/import_odoo.py --source examples/enset/stages.candidature.sejour.xlsx
        [--out examples/enset/dossier-u3.xlsx] [--cloture-precedente AAAA-MM-JJ]

Pré-remplit les feuilles Candidats (id, email — identifiant de connexion Odoo et de
l'application web —, nom, rang, département, destination, durée, montant d'indemnité)
et Historique (date du dernier stage). Contrôle au passage la
cohérence des zones et des montants d'indemnité calculés par Odoo avec notre
référentiel (data/costs).
Les critères/activités restent à collecter (fiche de déclaration ou formulaire).
"""

from __future__ import annotations

import argparse
import sys
from datetime import date, datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from classement.costs import indemnite, load_costs, zone_of
from classement.excel_io import write_template
from classement.grids import find_grid
from classement.institutions import load_institution

GRID_ID = "u3-residences-scientifiques"

DEPARTEMENTS = {
    "Math et Informatique": "Département de Mathématiques et Informatique",
    "Physique et Chimie": "Département de Physique et Chimie",
    "Sciences Naturelles": "Département des Sciences Naturelles",
    "Technologies": "Département de Technologie",
}

GRADES = {
    "Professeur": "Professeur",
    "Maître de conference classe (A)": "Maître de conférences A",
    "Maître de conference classe (B)": "Maître de conférences B",
    "Professeur émérite": "Professeur émérite",
}


def _date(value) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()[:10] if value else None


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", required=True)
    parser.add_argument("--out", default=str(ROOT / "examples" / "enset" / "dossier-u3.xlsx"))
    parser.add_argument(
        "--cloture-precedente",
        metavar="AAAA-MM-JJ",
        help="Date de clôture du dépôt de la dernière campagne, fixée par la commission "
        "et appliquée à toutes les lignes de l'Historique (colonne "
        "date_cloture_plateforme). Sans cette option la colonne reste vide et le "
        "moteur se rabat sur la date de mobilité.",
    )
    args = parser.parse_args()
    cloture = None
    if args.cloture_precedente:
        cloture = date.fromisoformat(args.cloture_precedente).isoformat()

    grid = find_grid(GRID_ID)
    enset = load_institution("enset-skikda")
    costs = load_costs()

    src = load_workbook(args.source, data_only=True)
    ws = src[src.sheetnames[0]]
    cols = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

    def cell(row: int, name: str):
        return ws.cell(row=row, column=cols[name]).value

    # classeur cible : modèle u3 vierge puis remplissage
    write_template(args.out, grid, enset)
    out = load_workbook(args.out)
    cand = out["Candidats"]
    hist = out["Historique"]
    headers = {cand.cell(row=1, column=c).value: c for c in range(1, cand.max_column + 1)}

    warnings: list[str] = []
    zone_mismatches: list[str] = []
    montant_mismatches: list[str] = []
    n, hist_row = 0, 2

    for row in range(2, ws.max_row + 1):
        ref = cell(row, "Référence")
        if not ref:
            continue
        n += 1
        target = n + 1  # ligne dans la feuille Candidats

        nom = str(cell(row, "Enseignant") or "").strip()
        dept_odoo = str(cell(row, "Département") or "").strip()
        dept = DEPARTEMENTS.get(dept_odoo)
        if dept is None:
            warnings.append(
                f"{ref} ({nom}) : département Odoo {dept_odoo!r} sans équivalent ENSET — "
                f"laissé vide, à corriger (hors champ u3 ?)."
            )
        grade_odoo = str(cell(row, "Grade scientifique") or "").strip()
        grade = GRADES.get(grade_odoo)
        if grade is None:
            warnings.append(f"{ref} ({nom}) : grade {grade_odoo!r} non reconnu — laissé vide.")

        pays = str(cell(row, "Pays") or "").strip()
        if pays.casefold() == "algérie".casefold():
            warnings.append(
                f"{ref} ({nom}) : destination « Algérie » — une résidence à l'étranger "
                f"ne peut pas se dérouler en Algérie, dossier à vérifier."
            )
        duree = cell(row, "Durée")

        cand.cell(row=target, column=headers["id"], value=str(ref))
        email = str(cell(row, "Email") or "").strip().lower() if "Email" in cols else ""
        if email:
            cand.cell(row=target, column=headers["email"], value=email)
        else:
            warnings.append(
                f"{ref} ({nom}) : e-mail absent de l'export Odoo — compte web à créer à la main."
            )
        cand.cell(row=target, column=headers["nom_prenom"], value=nom)
        cand.cell(row=target, column=headers["population"], value="enseignant_chercheur")
        if dept:
            cand.cell(row=target, column=headers["departement"], value=dept)
        if grade:
            cand.cell(row=target, column=headers["rang_scientifique"], value=grade)
        cand.cell(row=target, column=headers["pays_destination"], value=pays)
        if duree:
            cand.cell(row=target, column=headers["duree_jours"], value=int(duree))

        # historique : date du dernier stage connue d'Odoo
        last = _date(cell(row, "Date dernier Stage"))
        if last:
            hist.cell(row=hist_row, column=1, value=str(ref))
            hist.cell(row=hist_row, column=2, value=last)
            if cloture:
                hist.cell(row=hist_row, column=3, value=cloture)
            hist_row += 1

        # contrôles croisés avec les calculs d'Odoo
        zone_odoo = str(cell(row, "Zone") or "").replace(" ", "").casefold()  # "zone1"/"zone2"
        zone_calc, _ = zone_of(pays, costs)
        if zone_odoo and zone_odoo != zone_calc:
            zone_mismatches.append(f"{ref} ({nom}) : {pays} — Odoo {zone_odoo}, référentiel {zone_calc}")

        montant_odoo = cell(row, "Montant")
        if montant_odoo:
            cand.cell(
                row=target, column=headers["montant_indemnite (DA)"], value=float(montant_odoo)
            )
        if montant_odoo and duree:
            calc, _ = indemnite(costs, "residence_scientifique", zone_calc, int(duree), "enseignant_chercheur")
            if abs(float(montant_odoo) - calc) > 0.01:
                montant_mismatches.append(
                    f"{ref} ({nom}) : {pays} {duree} j — Odoo {float(montant_odoo):,.0f} DA, "
                    f"référentiel {calc:,.0f} DA"
                )

    out.save(args.out)
    print(f"{n} candidatures importées dans {args.out}")
    print(f"Historique : {hist_row - 2} date(s) de dernier stage reportée(s)."
          + (f" Clôture précédente appliquée : {cloture}." if cloture
             else " Clôture précédente non renseignée (repli sur la date de mobilité)."))
    for w in warnings:
        print(f"  ⚠ {w}")
    if zone_mismatches:
        print(f"-- {len(zone_mismatches)} écart(s) de zone Odoo / référentiel --")
        for m in zone_mismatches:
            print(f"  ⚠ {m}")
    if montant_mismatches:
        print(f"-- {len(montant_mismatches)} écart(s) de montant Odoo / référentiel --")
        for m in montant_mismatches:
            print(f"  ⚠ {m}")
    if not zone_mismatches and not montant_mismatches:
        print("Zones et montants Odoo cohérents avec le référentiel data/costs.")
    print("Reste à collecter : critères et activités (+ billet estimé et frais divers).")
    return 0


if __name__ == "__main__":
    sys.exit(main())
