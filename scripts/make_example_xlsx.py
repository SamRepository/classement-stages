"""Génère le modèle de saisie u4 pour l'ENSET et un classeur d'exemple rempli
(mêmes dossiers que examples/enset/candidats-u4.json).

Usage : python scripts/make_example_xlsx.py
"""

from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from classement.excel_io import write_template
from classement.grids import find_grid
from classement.institutions import load_institution

OUT_DIR = ROOT / "examples" / "enset"


def _col(ws, header: str) -> int:
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == header:
            return col
    raise KeyError(header)


def main() -> None:
    enset = load_institution("enset-skikda")

    # Modèles vierges pour les grilles applicables à l'ENSET (u1 et u3 incluent
    # la colonne rang_scientifique avec menu déroulant des grades en français).
    locked: list[str] = []
    for grid_id, name in [
        ("u1-manifestations-internationales", "modele-u1.xlsx"),
        ("u2-personnel-administratif", "modele-u2.xlsx"),
        ("u3-residences-scientifiques", "modele-u3.xlsx"),
        ("u4-perfectionnement", "modele-u4.xlsx"),
    ]:
        path = OUT_DIR / name
        try:
            write_template(path, find_grid(grid_id), enset)
            print(f"Modèle : {path}")
        except PermissionError:
            locked.append(name)
            print(f"IGNORÉ (ouvert dans Excel ?) : {path}")

    grid = find_grid("u4-perfectionnement")
    template_path = OUT_DIR / "modele-u4.xlsx"
    if "modele-u4.xlsx" in locked or not template_path.exists():
        print("Exemple rempli non régénéré : le modèle u4 n'a pas pu être réécrit.")
        return

    wb = load_workbook(template_path)
    cand = wb["Candidats"]
    act = wb["Activites"]
    hist = wb["Historique"]

    rows = [
        # id, nom, prenom, population, departement(label), inscriptions, startup,
        # attestations, pays, duree_jours, billet (DA), frais divers (DA)
        ("D101", "Bouzid", "Sara", "doctorant_non_salarie", "Département de Technologie",
         3, "Oui", None, "France", 21, 95000, 28000),
        ("D102", "Khelifi", "Mohamed", "doctorant_non_salarie", "Département de Technologie",
         4, "Non", 1, "Tunisie", 30, 38000, 12000),
        ("D103", "Aissaoui", "Rym", "doctorant_non_salarie",
         "Département de Mathématiques et Informatique", 2, "Non", None,
         "Espagne", 15, 88000, 26000),
        ("E201", "Ziani", "Hocine", "maitre_assistant", "Département de Physique et Chimie",
         None, "Non", None, "Japon", 15, 460000, 35000),
    ]
    columns = {
        "id": _col(cand, "id"),
        "nom": _col(cand, "nom_prenom"),
        "population": _col(cand, "population"),
        "departement": _col(cand, "departement"),
        "pays": _col(cand, "pays_destination"),
        "duree": _col(cand, "duree_jours"),
        "billet": _col(cand, "billet_estime (DA)"),
        "frais": _col(cand, "frais_divers (DA)"),
        "inscription": _col(cand, "inscription_doctorat.inscription (qte)"),
        "startup": _col(cand, "projet_startup_incubateur (Oui/Non)"),
        "attestation": _col(cand, "structures_accompagnement.attestation (qte)"),
        "polycopie": _col(cand, "polycopie_pedagogique (Oui/Non)"),
        "polycopie_b1": _col(cand, "polycopie_pedagogique (bonus 1 Oui/Non)"),
    }
    for r, row in enumerate(rows, start=2):
        cid, nom, prenom, pop, dept, insc, startup, attest, pays, duree, billet, frais = row
        cand.cell(row=r, column=columns["id"], value=cid)
        cand.cell(row=r, column=columns["nom"], value=f"{nom} {prenom}")
        cand.cell(row=r, column=columns["population"], value=pop)
        cand.cell(row=r, column=columns["departement"], value=dept)
        cand.cell(row=r, column=columns["pays"], value=pays)
        cand.cell(row=r, column=columns["duree"], value=duree)
        cand.cell(row=r, column=columns["billet"], value=billet)
        cand.cell(row=r, column=columns["frais"], value=frais)
        if insc:
            cand.cell(row=r, column=columns["inscription"], value=insc)
        cand.cell(row=r, column=columns["startup"], value=startup)
        if attest:
            cand.cell(row=r, column=columns["attestation"], value=attest)
    # E201 : polycopié approuvé + bonus anglais
    cand.cell(row=5, column=columns["polycopie"], value="Oui")
    cand.cell(row=5, column=columns["polycopie_b1"], value="Oui")

    act_headers = {act.cell(row=1, column=c).value: c for c in range(1, act.max_column + 1)}
    activities = [
        # candidat, element, qte, position, date, doi, url, porteur, bonus
        ("D101", "publications :: classe_b", 1, 1, "2025-11-20",
         "10.1016/j.example.2025.11.001", None, None, None),
        ("D101", "communications :: intl_indexee_scopus_wos", 1, None, "2025-04-18",
         None, "https://www.scopus.com/record/display.uri?eid=2-s2.0-EXEMPLE", None, None),
        ("D101", "communications :: nationale", 2, None, "2024-12-05", None, None, None, None),
        ("D102", "publications :: classe_a", 1, 2, "2025-09-01",
         "10.1007/s00example-2025-0042", None, None, None),
        ("D102", "publications :: classe_c", 1, 1, "2024-06-10",
         None, "https://www.asjp.cerist.dz/en/article/EXEMPLE", None, None),
        ("D103", "communications :: intl_referencee", 2, None, "2025-10-14", None, None, None, None),
        ("E201", "elearning :: cours", 2, None, None, None, None, None, 1),
        ("E201", "elearning :: tp", 2, None, None, None, None, None, None),
        ("E201", "encadrement_master :: memoire_master", 2, None, "2025-06-30", None, None, None, None),
        ("E201", "publications :: classe_c", 1, 1, "2025-02-15",
         "10.5281/zenodo.EXEMPLE", None, None, None),
    ]
    columns_order = ["candidat_id", "element", "quantite", "position_auteur",
                     "date (AAAA-MM-JJ)", "doi", "url", "porteur_nb", "bonus_nb"]
    for r, row in enumerate(activities, start=2):
        for header, value in zip(columns_order, row):
            if value is not None:
                act.cell(row=r, column=act_headers[header], value=value)

    hist.cell(row=2, column=1, value="D102")
    hist.cell(row=2, column=2, value="2025-02-10")
    hist.cell(row=2, column=3, value="2024-11-30")

    filled_path = OUT_DIR / "dossiers-u4.xlsx"
    try:
        wb.save(filled_path)
        print(f"Exemple rempli : {filled_path}")
    except PermissionError:
        print(f"IGNORÉ (ouvert dans Excel ?) : {filled_path}")


if __name__ == "__main__":
    main()
