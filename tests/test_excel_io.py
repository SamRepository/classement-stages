"""Tests du modèle de saisie Excel, de l'import et des exports officiels."""

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from classement.engine import score_candidate
from classement.excel_io import build_template, column_plan, read_candidates, write_template
from classement.exports import export_fiches, export_html, export_pv
from classement.grids import find_grid, load_shared_rules
from classement.institutions import group_by_for, load_institution
from classement.ranking import rank_candidates

ROOT = Path(__file__).resolve().parent.parent


@pytest.fixture(scope="module")
def u4():
    return find_grid("u4-perfectionnement")


@pytest.fixture(scope="module")
def u2():
    return find_grid("u2-personnel-administratif")


@pytest.fixture(scope="module")
def enset():
    return load_institution("enset-skikda")


def test_column_plan_routes_counts(u4):
    columns, activities = column_plan(u4)
    headers = [c["header"] for c in columns]
    # count simple (pas de fenêtre/pondération) -> colonne Candidats
    assert "inscription_doctorat.inscription (qte)" in headers
    # count avec fenêtre ou pondération -> feuille Activites
    activity_keys = {a["key"] for a in activities}
    assert "publications :: classe_a" in activity_keys
    assert "elearning :: cours" in activity_keys
    assert not any(h.startswith("publications.") for h in headers)


def test_template_sheets_and_dropdowns(u4, enset, tmp_path):
    path = tmp_path / "modele.xlsx"
    write_template(path, u4, enset)
    wb = load_workbook(path)
    assert set(wb.sheetnames) >= {"Candidats", "Activites", "Historique", "Listes", "Referentiel"}
    listes = wb["Listes"]
    values = [listes.cell(row=r, column=c).value for c in range(1, listes.max_column + 1)
              for r in range(1, listes.max_row + 1)]
    assert "Département de Technologie" in values
    assert "doctorant_non_salarie" in values


def test_u2_template_has_no_activities_sheet_entries(u2):
    _columns, activities = column_plan(u2)
    assert activities == []  # tout u2 se saisit dans la feuille Candidats


def test_u3_template_rank_dropdown_uses_french_labels(enset, tmp_path):
    u3 = find_grid("u3-residences-scientifiques")
    path = tmp_path / "modele-u3.xlsx"
    write_template(path, u3, enset)
    wb = load_workbook(path)

    cand = wb["Candidats"]
    headers = [cand.cell(row=1, column=c).value for c in range(1, cand.max_column + 1)]
    assert "rang_scientifique" in headers
    assert "rang_scientifique (bonus Oui/Non)" in headers  # bonus habilitation MCB

    listes = wb["Listes"]
    values = [listes.cell(row=r, column=c).value for c in range(1, listes.max_column + 1)
              for r in range(2, listes.max_row + 1)]
    assert "Professeur émérite" in values
    assert "Maître de conférences B" in values
    assert "professeur_emerite" not in values  # libellés affichés, pas les ids


def test_u3_citations_simple_entry_with_profile_url(enset, tmp_path):
    u3 = find_grid("u3-residences-scientifiques")
    path = tmp_path / "citations.xlsx"
    write_template(path, u3, enset)
    wb = load_workbook(path)
    cand = wb["Candidats"]
    headers = {cand.cell(row=1, column=c).value: c for c in range(1, cand.max_column + 1)}
    # nom et prénom fusionnés ; citations en saisie simple (qte + url profil)
    assert "nom_prenom" in headers
    assert "nom" not in headers and "prenom" not in headers
    assert "citations_scopus.citation (qte)" in headers
    assert "citations_scopus (url profil)" in headers

    cand.cell(row=2, column=headers["id"], value="C1")
    cand.cell(row=2, column=headers["nom_prenom"], value="BABOURI LAIDI")
    cand.cell(row=2, column=headers["citations_scopus.citation (qte)"], value=120)
    cand.cell(row=2, column=headers["citations_scopus (url profil)"],
              value="https://scholar.google.com/citations?user=EXEMPLE")
    wb.save(path)

    candidates, errors = read_candidates(path, u3, enset)
    assert errors == []
    assert candidates[0]["nom"] == "BABOURI LAIDI"
    item = candidates[0]["entries"]["citations_scopus"]["items"][0]
    assert item["count"] == 120
    assert "scholar.google.com" in item["url"]

    shared = load_shared_rules()
    breakdown = score_candidate(u3, candidates[0], shared, "2026-06-30")
    line = next(l for l in breakdown.lines if l.criterion_id == "citations_scopus")
    assert line.points == pytest.approx(12.0)  # 120 × 0,1
    assert any("réf." in d and "scholar" in d for d in line.details)


def test_enum_import_accepts_label_or_value(enset, tmp_path):
    u3 = find_grid("u3-residences-scientifiques")
    path = tmp_path / "saisie-u3.xlsx"
    write_template(path, u3, enset)
    wb = load_workbook(path)
    cand = wb["Candidats"]
    headers = {cand.cell(row=1, column=c).value: c for c in range(1, cand.max_column + 1)}
    # candidat 1 : libellé français ; candidat 2 : valeur brute ; candidat 3 : invalide
    rows = [("P1", "Maître de conférences B", "Oui"), ("P2", "professeur", None), ("P3", "Recteur", None)]
    for r, (cid, rang, bonus) in enumerate(rows, start=2):
        cand.cell(row=r, column=headers["id"], value=cid)
        cand.cell(row=r, column=headers["rang_scientifique"], value=rang)
        if bonus:
            cand.cell(row=r, column=headers["rang_scientifique (bonus Oui/Non)"], value=bonus)
    wb.save(path)

    candidates, errors = read_candidates(path, u3, enset)
    by_id = {c["id"]: c for c in candidates}
    assert by_id["P1"]["entries"]["rang_scientifique"] == {"value": "mcb", "option_bonus": True}
    assert by_id["P2"]["entries"]["rang_scientifique"]["value"] == "professeur"
    assert any("Recteur" in e and "hors barème" in e for e in errors)

    shared = load_shared_rules()
    assert score_candidate(u3, by_id["P1"], shared, "2026-06-30").lines[0].points == 7  # 3 + 4


def test_roundtrip_excel_matches_json_example(u4, enset):
    xlsx = ROOT / "examples" / "enset" / "dossiers-u4.xlsx"
    if not xlsx.exists():
        pytest.skip("exemple non généré (scripts/make_example_xlsx.py)")
    with open(ROOT / "examples" / "enset" / "candidats-u4.json", encoding="utf-8") as fh:
        json_candidates = json.load(fh)["candidates"]

    excel_candidates, errors = read_candidates(xlsx, u4, enset)
    assert errors == []
    shared = load_shared_rules()
    totals_json = {
        c["id"]: score_candidate(u4, c, shared, "2026-06-30").total for c in json_candidates
    }
    totals_excel = {
        c["id"]: score_candidate(u4, c, shared, "2026-06-30").total for c in excel_candidates
    }
    assert totals_excel == totals_json


def test_import_reports_errors(u4, enset, tmp_path):
    path = tmp_path / "errs.xlsx"
    write_template(path, u4, enset)
    wb = load_workbook(path)
    cand = wb["Candidats"]
    cand.cell(row=2, column=1, value="X1")
    # département inconnu + valeur Oui/Non invalide
    headers = {cand.cell(row=1, column=c).value: c for c in range(1, cand.max_column + 1)}
    cand.cell(row=2, column=headers["departement"], value="Génie Civil")
    cand.cell(row=2, column=headers["projet_startup_incubateur (Oui/Non)"], value="peut-être")
    act = wb["Activites"]
    act.cell(row=2, column=1, value="INCONNU")
    act.cell(row=2, column=2, value="publications :: classe_a")
    wb.save(path)

    candidates, errors = read_candidates(path, u4, enset)
    assert len(candidates) == 1
    assert any("département inconnu" in e.lower() for e in errors)
    assert any("Oui/Non invalide" in e for e in errors)
    assert any("candidat inconnu" in e for e in errors)


def test_activities_doi_url_roundtrip_and_reference_warning(u4, enset, tmp_path):
    path = tmp_path / "refs.xlsx"
    write_template(path, u4, enset)
    wb = load_workbook(path)
    cand = wb["Candidats"]
    cand.cell(row=2, column=1, value="R1")
    act = wb["Activites"]
    headers = {act.cell(row=1, column=c).value: c for c in range(1, act.max_column + 1)}
    assert "doi" in headers and "url" in headers
    # publication avec DOI, et une seconde sans référence (doit produire un avertissement)
    act.cell(row=2, column=headers["candidat_id"], value="R1")
    act.cell(row=2, column=headers["element"], value="publications :: classe_a")
    act.cell(row=2, column=headers["position_auteur"], value=1)
    act.cell(row=2, column=headers["date (AAAA-MM-JJ)"], value="2025-05-01")
    act.cell(row=2, column=headers["doi"], value="10.1234/abc")
    act.cell(row=3, column=headers["candidat_id"], value="R1")
    act.cell(row=3, column=headers["element"], value="publications :: classe_b")
    act.cell(row=3, column=headers["position_auteur"], value=1)
    act.cell(row=3, column=headers["date (AAAA-MM-JJ)"], value="2025-06-01")
    wb.save(path)

    candidates, errors = read_candidates(path, u4, enset)
    assert errors == []
    items = candidates[0]["entries"]["publications"]["items"]
    assert {"doi": "10.1234/abc"}.items() <= items[0].items()

    shared = load_shared_rules()
    breakdown = score_candidate(u4, candidates[0], shared, "2026-06-30")
    pubs = next(l for l in breakdown.lines if l.criterion_id == "publications")
    assert any("réf. 10.1234/abc" in d for d in pubs.details)
    assert any("classe_b" in w and "DOI/URL non fourni" in w for w in pubs.warnings)


def test_old_workbook_without_doi_columns_still_imports(u4, enset, tmp_path):
    path = tmp_path / "ancien.xlsx"
    write_template(path, u4, enset)
    wb = load_workbook(path)
    act = wb["Activites"]
    # simule un ancien classeur : colonnes doi/url supprimées
    headers = {act.cell(row=1, column=c).value: c for c in range(1, act.max_column + 1)}
    act.delete_cols(headers["doi"], 2)
    cand = wb["Candidats"]
    cand.cell(row=2, column=1, value="A1")
    act.cell(row=2, column=1, value="A1")
    act.cell(row=2, column=2, value="communications :: nationale")
    act.cell(row=2, column=3, value=2)
    wb.save(path)

    candidates, errors = read_candidates(path, u4, enset)
    assert errors == []
    assert candidates[0]["entries"]["communications"]["items"][0]["count"] == 2


def test_write_candidates_roundtrip(u4, enset, tmp_path):
    from classement.excel_io import write_candidates

    with open(ROOT / "examples" / "enset" / "candidats-u4.json", encoding="utf-8") as fh:
        source = json.load(fh)["candidates"]
    path = tmp_path / "ecrit.xlsx"
    write_candidates(path, u4, enset, source)
    reread, errors = read_candidates(path, u4, enset)
    assert errors == []
    assert len(reread) == len(source)

    shared = load_shared_rules()
    for original, copy in zip(source, sorted(reread, key=lambda c: [s["id"] for s in source].index(c["id"]))):
        a = score_candidate(u4, original, shared, "2026-06-30").total
        b = score_candidate(u4, copy, shared, "2026-06-30").total
        assert b == pytest.approx(a), original["id"]


def test_exports_pv_fiches_html(u4, enset, tmp_path):
    xlsx = ROOT / "examples" / "enset" / "dossiers-u4.xlsx"
    if not xlsx.exists():
        pytest.skip("exemple non généré")
    candidates, _ = read_candidates(xlsx, u4, enset)
    shared = load_shared_rules()
    breakdowns = [score_candidate(u4, c, shared, "2026-06-30") for c in candidates]
    resolver = lambda c: group_by_for(enset, u4["id"], c.get("population"))  # noqa: E731
    groups = rank_candidates(candidates, breakdowns, group_by=resolver)

    pv = tmp_path / "pv.xlsx"
    export_pv(pv, groups, candidates, u4, enset, "2026-06-30")
    wb = load_workbook(pv)
    assert len(wb.sheetnames) == len(groups)
    first = wb[wb.sheetnames[0]]
    text = [cell.value for row in first.iter_rows() for cell in row if cell.value]
    assert any("École Normale Supérieure" in str(v) for v in text)
    assert "Rang" in text

    fiches = tmp_path / "fiches.xlsx"
    export_fiches(fiches, breakdowns, candidates, groups, u4, enset, "2026-06-30")
    wb = load_workbook(fiches)
    assert set(wb.sheetnames) == {b.candidate_id for b in breakdowns}
    d101 = wb["D101"]
    values = [cell.value for row in d101.iter_rows() for cell in row if cell.value is not None]
    assert "TOTAL" in values
    assert 28 in values or 28.0 in values

    html_path = tmp_path / "doc.html"
    export_html(html_path, groups, breakdowns, candidates, u4, enset, "2026-06-30")
    content = html_path.read_text(encoding="utf-8")
    assert "Fiche d'évaluation — Bouzid Sara (D101)" in content
    assert "École Normale Supérieure" in content
