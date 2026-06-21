"""Espace admin : comptes, import, bénéfices, campagne, réouverture."""

import io

from openpyxl import Workbook
from sqlalchemy import select

from tests.webapp.conftest import login
from webapp.models import Benefit, Dossier, User


def test_creation_compte_et_connexion(client, db_session, campaign, admin):
    login(client, "admin@test.dz")
    r = client.post("/admin/utilisateurs",
                    data={"email": "nouveau@test.dz", "nom": "Mansouri", "prenom": "B",
                          "role": "enseignant"})
    assert r.status_code == 200
    assert "nouveau@test.dz" in r.text
    # Le mot de passe initial apparaît une seule fois dans la réponse.
    nouveau = db_session.scalar(select(User).where(User.email == "nouveau@test.dz"))
    assert nouveau.role == "enseignant"


def test_liste_comptes_triable(client, db_session, campaign, admin):
    db_session.add_all([
        User(email="zoe@test.dz", password_hash="x", nom="Zoe", role="enseignant"),
        User(email="amir@test.dz", password_hash="x", nom="Amir", role="commission"),
    ])
    db_session.commit()
    login(client, "admin@test.dz")
    # Les en-têtes proposent le tri par email et par rôle.
    page = client.get("/admin/utilisateurs")
    assert "/admin/utilisateurs?tri=email" in page.text
    assert "/admin/utilisateurs?tri=role" in page.text
    # Tri par email : amir@ apparaît avant zoe@ dans le HTML.
    par_email = client.get("/admin/utilisateurs?tri=email").text
    assert par_email.index("amir@test.dz") < par_email.index("zoe@test.dz")


def test_import_xlsx_cree_comptes_et_dossiers(client, db_session, campaign, admin):
    wb = Workbook()
    ws = wb.active
    ws.append(["Email", "Nom", "Prénom", "Référence", "Département"])
    ws.append(["a@test.dz", "Alpha", "A", "DC-2026-101", "technologie"])
    ws.append(["b@test.dz", "Beta", "B", "DC-2026-102", ""])
    ws.append(["a@test.dz", "Doublon", "", "", ""])  # déjà créé ligne 2
    buffer = io.BytesIO()
    wb.save(buffer)

    login(client, "admin@test.dz")
    r = client.post("/admin/utilisateurs/import",
                    files={"fichier": ("comptes.xlsx", buffer.getvalue(),
                                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 200
    assert "a@test.dz" in r.text and "b@test.dz" in r.text
    assert "existe déjà" in r.text

    refs = set(db_session.scalars(select(Dossier.candidate_ref)))
    assert {"DC-2026-101", "DC-2026-102"} <= refs


def _classeur_u3() -> bytes:
    """Mini dossier-u3.xlsx : feuille Candidats (avec mobilité) + feuille Historique."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidats"
    ws.append(["id", "email", "nom_prenom", "population", "departement",
               "pays_destination", "duree_jours", "billet_estime (DA)"])
    ws.append(["DC/2026/264", "candidat@univ.dz", "MERIMECHE IMENE",
               "enseignant_chercheur", "Département de Technologie",
               "Espagne", 10, 92000])
    hist = wb.create_sheet("Historique")
    hist.append(["candidat_id", "date_mobilite (AAAA-MM-JJ)", "date_cloture_plateforme (AAAA-MM-JJ)"])
    hist.append(["DC/2026/264", "2024-09-15", "2024-04-30"])
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_import_feuille_candidats_circuit_excel(client, db_session, campaign, admin):
    """Le dossier u3 issu d'Odoo s'importe tel quel : email = identifiant de
    connexion (comme Odoo), libellé de département converti en id du profil,
    mobilité reprise sur le dossier, feuille Historique → table benefits."""
    from webapp.models import Benefit

    login(client, "admin@test.dz")
    contenu = _classeur_u3()
    r = client.post("/admin/utilisateurs/import",
                    files={"fichier": ("dossier-u3.xlsx", contenu,
                                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 200
    user = db_session.scalar(select(User).where(User.email == "candidat@univ.dz"))
    assert user is not None and user.nom == "MERIMECHE IMENE"
    dossier = db_session.scalar(select(Dossier).where(Dossier.user_id == user.id))
    assert dossier.candidate_ref == "DC/2026/264"
    assert dossier.departement == "technologie"  # libellé converti en id du profil
    assert dossier.pays == "Espagne"
    assert dossier.duree_jours == 10
    assert dossier.billet_estime_da == 92000.0
    benefit = db_session.scalar(select(Benefit).where(Benefit.user_id == user.id))
    assert benefit is not None
    assert benefit.date.isoformat() == "2024-09-15"
    assert benefit.platform_close_date.isoformat() == "2024-04-30"

    # Réimport : ni compte ni bénéfice dupliqués.
    r = client.post("/admin/utilisateurs/import",
                    files={"fichier": ("dossier-u3.xlsx", contenu,
                                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 200
    assert "existe déjà" in r.text
    benefices = list(db_session.scalars(select(Benefit).where(Benefit.user_id == user.id)))
    assert len(benefices) == 1


def test_benefices_ajout_suppression(client, db_session, campaign, admin, enseignant):
    login(client, "admin@test.dz")
    r = client.post("/admin/benefices",
                    data={"user_id": str(enseignant.id), "date": "2024-09-15",
                          "platform_close_date": "2024-04-30", "note": "stage 2024"})
    assert r.status_code == 303
    benefit = db_session.scalar(select(Benefit))
    assert benefit.user_id == enseignant.id
    r = client.post(f"/admin/benefices/{benefit.id}/supprimer")
    assert r.status_code == 303
    assert db_session.scalar(select(Benefit)) is None


def test_import_benefices_par_email(client, db_session, campaign, admin, enseignant):
    login(client, "admin@test.dz")
    csv = b"email,date,cloture\nenseignant@test.dz,2024-05-10,2024-06-30\n"
    r = client.post("/admin/benefices/import",
                    files={"fichier": ("hist.csv", csv, "text/csv")})
    assert r.status_code == 200
    assert "1 b" in r.text  # « 1 bénéfice(s) importé(s) »
    benefits = list(db_session.scalars(select(Benefit).where(Benefit.user_id == enseignant.id)))
    assert len(benefits) == 1
    assert benefits[0].date.isoformat() == "2024-05-10"
    assert benefits[0].platform_close_date.isoformat() == "2024-06-30"
    # Idempotent : un réimport ne duplique pas.
    r = client.post("/admin/benefices/import",
                    files={"fichier": ("hist.csv", csv, "text/csv")})
    assert "0 b" in r.text
    assert len(list(db_session.scalars(select(Benefit)))) == 1


def test_import_benefices_email_inconnu_ignore(client, db_session, campaign, admin):
    login(client, "admin@test.dz")
    csv = b"email,date\ninconnu@test.dz,2024-05-10\n"
    r = client.post("/admin/benefices/import",
                    files={"fichier": ("h.csv", csv, "text/csv")})
    assert r.status_code == 200
    assert "0 b" in r.text
    assert "aucun compte" in r.text
    assert db_session.scalar(select(Benefit)) is None


def test_campagne_fenetre_intervalle(client, db_session, campaign, admin):
    login(client, "admin@test.dz")
    r = client.post("/admin/campagne", data={
        "statut": "ouverte", "campaign_date": "2026-06-30",
        "window_reference": "mobilite",
        "window_start_date": "2025-01-01", "window_end_date": "2025-12-31",
    })
    assert r.status_code == 303
    db_session.expire_all()
    assert campaign.window_reference == "mobilite"
    assert campaign.window_start_date.isoformat() == "2025-01-01"
    assert campaign.window_end_date.isoformat() == "2025-12-31"


def test_campagne_fenetre_intervalle_invalide(client, db_session, campaign, admin):
    """Début postérieur à la fin : rejet 422, paramètres inchangés."""
    login(client, "admin@test.dz")
    r = client.post("/admin/campagne", data={
        "statut": "ouverte", "campaign_date": "2026-06-30",
        "window_start_date": "2025-12-31", "window_end_date": "2025-01-01",
    })
    assert r.status_code == 422


def test_cloture_campagne(client, db_session, campaign, admin):
    login(client, "admin@test.dz")
    r = client.post("/admin/campagne", data={"statut": "cloturee",
                                             "campaign_date": "2026-06-30"})
    assert r.status_code == 303
    db_session.expire_all()
    assert campaign.statut == "cloturee"


def test_cloture_saisie_auto_soumet_brouillons(client, db_session, campaign, dossier, admin):
    """À la fermeture de la saisie, les dossiers en brouillon sont soumis d'office."""
    assert dossier.statut == "brouillon"
    login(client, "admin@test.dz")
    r = client.post("/admin/campagne", data={"statut": "cloturee",
                                             "campaign_date": "2026-06-30"})
    assert r.status_code == 303
    db_session.expire_all()
    assert dossier.statut == "soumis"
    assert dossier.submitted_at is not None


def test_admin_budget_saisie(client, db_session, campaign, dossier, admin):
    login(client, "admin@test.dz")
    r = client.post(f"/admin/budget/{dossier.id}",
                    data={"billet_estime_da": "85000", "frais_divers_da": "12000,50"})
    assert r.status_code == 303
    db_session.expire_all()
    assert dossier.billet_estime_da == 85000.0
    assert dossier.frais_divers_da == 12000.5


def test_reouverture_dossier(client, db_session, campaign, dossier, admin):
    dossier.statut = "soumis"
    db_session.commit()
    login(client, "admin@test.dz")
    r = client.post(f"/admin/dossiers/{dossier.id}/reouvrir")
    assert r.status_code == 303
    db_session.expire_all()
    assert dossier.statut == "brouillon"


def test_admin_interdit_aux_autres_roles(client, campaign, enseignant, membre_commission):
    login(client, "enseignant@test.dz")
    assert client.get("/admin/utilisateurs").status_code == 403
    client.post("/deconnexion")
    login(client, "commission@test.dz")
    assert client.get("/admin/campagne").status_code == 403
