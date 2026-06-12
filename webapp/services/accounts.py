"""Création et import des comptes enseignants (CSV ou Excel).

Source privilégiée : le classeur `dossier-u3.xlsx` produit par
`scripts/import_odoo.py` — la feuille Candidats fournit comptes et mobilité
(pays, durée, billet, frais), la feuille Historique alimente la table des
bénéfices (pénalité « 3 − n » et fenêtre « après dernier bénéfice »).
Un CSV simple (email, nom, …) reste accepté.

Colonnes reconnues (en-têtes insensibles à la casse et aux accents, repérage
par mot-clé) : email, nom/nom_prenom, prénom, référence (« ref »/« id »),
département, pays, durée, billet, frais.
"""

from __future__ import annotations

import csv
import io
import unicodedata
from datetime import date, datetime

from fastapi import HTTPException
from sqlalchemy import select
from sqlalchemy.orm import Session

from webapp.models import Benefit, Campaign, Dossier, User
from webapp.security import generate_password, hash_password
from webapp.services.scoring import get_institution


def _normalize(text: str) -> str:
    t = unicodedata.normalize("NFKD", text.strip().lower())
    return "".join(c for c in t if not unicodedata.combining(c))


def _departement_resolver(institution_id: str):
    """id ou libellé de département (normalisé) → id du profil d'établissement."""
    table: dict[str, str] = {}
    for d in get_institution(institution_id).get("departements", []):
        table[_normalize(d["id"])] = d["id"]
        if d.get("label_fr"):
            table[_normalize(d["label_fr"])] = d["id"]

    def resolve(value: str) -> str | None:
        return table.get(_normalize(value)) if value else None

    return resolve


def _map_headers(headers: list[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for i, raw in enumerate(headers):
        h = unicodedata.normalize("NFKD", (raw or "").strip().lower())
        h = "".join(c for c in h if not unicodedata.combining(c))
        if "mail" in h and "email" not in mapping:
            mapping["email"] = i
        elif h == "nom_prenom" and "nom" not in mapping:
            mapping["nom"] = i  # nom complet (feuille Candidats du circuit Excel)
        elif "prenom" in h and "prenom" not in mapping:
            mapping["prenom"] = i
        elif "nom" in h and "nom" not in mapping:
            mapping["nom"] = i
        elif ("ref" in h or h == "id") and "ref" not in mapping:
            mapping["ref"] = i
        elif "depart" in h and "departement" not in mapping:
            mapping["departement"] = i
        elif "pays" in h and "pays" not in mapping:
            mapping["pays"] = i
        elif "duree" in h and "duree" not in mapping:
            mapping["duree"] = i
        elif "billet" in h and "billet" not in mapping:
            mapping["billet"] = i
        elif "frais" in h and "frais" not in mapping:
            mapping["frais"] = i
    if "email" not in mapping or "nom" not in mapping:
        raise HTTPException(
            status_code=422,
            detail="Colonnes « email » et « nom » introuvables dans le fichier.",
        )
    return mapping


def _rows_from_file(filename: str, content: bytes) -> tuple[list[list[str]], list[list]]:
    """(lignes Candidats en texte, lignes Historique brutes — vide pour un CSV)."""
    name = (filename or "").lower()
    if name.endswith(".xlsx"):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb["Candidats"] if "Candidats" in wb.sheetnames else wb.active
        rows = [
            [("" if c is None else str(c)) for c in row] for row in ws.iter_rows(values_only=True)
        ]
        history: list[list] = []
        if "Historique" in wb.sheetnames:
            history = [list(row) for row in wb["Historique"].iter_rows(values_only=True)]
        return rows, history
    if name.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        dialect = csv.Sniffer().sniff(text.splitlines()[0], delimiters=";,\t")
        return [list(r) for r in csv.reader(io.StringIO(text), dialect)], []
    raise HTTPException(status_code=422, detail="Format attendu : .csv ou .xlsx.")


def _coerce_date(value) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value).strip()[:10])
    except ValueError:
        return None


def _coerce_number(value) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(str(value).replace(",", ".").strip())
    except ValueError:
        return None


def import_accounts(
    db: Session, campaign: Campaign, filename: str, content: bytes
) -> tuple[list[dict], list[str]]:
    """Crée les comptes (rôle enseignant) et leur dossier brouillon.

    Retourne (créés : [{email, nom, prenom, password}], ignorés : messages).
    Les comptes existants sont conservés tels quels (import idempotent).
    """
    rows, history = _rows_from_file(filename, content)
    if not rows:
        raise HTTPException(status_code=422, detail="Fichier vide.")
    mapping = _map_headers(rows[0])
    resolve_departement = _departement_resolver(campaign.institution_id)

    created: list[dict] = []
    skipped: list[str] = []
    for index, row in enumerate(rows[1:], start=2):
        def cell(key: str) -> str:
            i = mapping.get(key)
            return (row[i] or "").strip() if i is not None and i < len(row) else ""

        email = cell("email").lower()
        nom = cell("nom")
        if not email or "@" not in email:
            if any((c or "").strip() for c in row):
                skipped.append(f"Ligne {index} : adresse électronique absente ou invalide.")
            continue
        if db.scalar(select(User).where(User.email == email)):
            skipped.append(f"Ligne {index} : {email} existe déjà (conservé).")
            continue
        password = generate_password()
        user = User(email=email, password_hash=hash_password(password),
                    nom=nom or email.split("@")[0], prenom=cell("prenom"), role="enseignant")
        db.add(user)
        db.flush()
        departement = resolve_departement(cell("departement"))
        if cell("departement") and departement is None:
            skipped.append(
                f"Ligne {index} : département {cell('departement')!r} inconnu du profil — "
                "laissé vide sur le dossier."
            )
        duree = _coerce_number(cell("duree"))
        db.add(Dossier(
            campaign_id=campaign.id,
            user_id=user.id,
            candidate_ref=cell("ref") or f"WEB-{user.id:03d}",
            departement=departement,
            pays=cell("pays") or None,
            duree_jours=int(duree) if duree else None,
            billet_estime_da=_coerce_number(cell("billet")),
            frais_divers_da=_coerce_number(cell("frais")),
        ))
        created.append({"email": email, "nom": user.nom, "prenom": user.prenom,
                        "password": password})
    db.flush()
    n_benefits = _import_benefits(db, campaign, history, skipped)
    if n_benefits:
        skipped.append(f"Historique : {n_benefits} bénéfice(s) antérieur(s) importé(s).")
    db.commit()
    return created, skipped


def _import_benefits(
    db: Session, campaign: Campaign, history: list[list], messages: list[str]
) -> int:
    """Feuille Historique (candidat_id, date_mobilite, date_cloture) → table benefits.

    Idempotent : un bénéfice déjà présent (même enseignant, même date) n'est pas
    dupliqué en cas de réimport.
    """
    if len(history) < 2:
        return 0
    header = [_normalize(str(h or "")) for h in history[0]]
    idx = {"ref": None, "date": None, "cloture": None}
    for i, h in enumerate(header):
        if "candidat" in h and idx["ref"] is None:
            idx["ref"] = i
        elif "cloture" in h and idx["cloture"] is None:
            idx["cloture"] = i
        elif "date" in h and idx["date"] is None:
            idx["date"] = i
    if idx["ref"] is None or idx["date"] is None:
        messages.append("Feuille Historique : en-têtes non reconnus, bénéfices non importés.")
        return 0

    def cell(row: list, key: str):
        i = idx[key]
        return row[i] if i is not None and i < len(row) else None

    count = 0
    for index, row in enumerate(history[1:], start=2):
        ref = str(cell(row, "ref") or "").strip()
        if not ref:
            continue
        dossier = db.scalar(
            select(Dossier).where(
                Dossier.campaign_id == campaign.id, Dossier.candidate_ref == ref
            )
        )
        if dossier is None:
            messages.append(f"Historique ligne {index} : référence {ref!r} sans dossier, ignorée.")
            continue
        date_mobilite = _coerce_date(cell(row, "date"))
        if date_mobilite is None:
            messages.append(f"Historique ligne {index} : date invalide pour {ref!r}, ignorée.")
            continue
        deja = db.scalar(
            select(Benefit).where(
                Benefit.user_id == dossier.user_id, Benefit.date == date_mobilite
            )
        )
        if deja:
            continue
        db.add(Benefit(
            user_id=dossier.user_id,
            date=date_mobilite,
            platform_close_date=_coerce_date(cell(row, "cloture")),
            source="import",
            note=f"Import {ref}",
        ))
        count += 1
    return count
